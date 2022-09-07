import pandas as pd
import numpy as np
import openpyxl
# Draw a new windows to reques the file yo get all data
from tkinter import *     # from tkinter import Tk for Python 3.x
from tkinter.filedialog import askopenfilename

# import module to print prettier
import pprintpp

# Import json library
import json


Tk().withdraw() # we don't want a full GUI, so keep the root window from appearing
filename = askopenfilename() # show an "Open" dialog box and return the path to the selected file
file = filename


data = pd.read_excel(file, sheet_name = 0, index_col=0)

# First function in start program
def index():
    sheets = getsheetname(file)
    converttolist(sheets)
    printf(sheets)

def getsheetname(file):
    sheets  = openpyxl.load_workbook(file)
    return sheets
    

def printf(sheets):
    pprintpp.pprint(sheets.sheetnames)


def converttolist(data):
    listResult = list(data.sheetnames)
    return listResult
index()


# data = pd.read_excel(file,sheet_name = "", index_col=0)
